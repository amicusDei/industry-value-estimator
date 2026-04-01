"""CSV and Excel export endpoints.

Supports two export formats:
- CSV: flat file with forecast data (single table)
- Excel: multi-sheet workbook with Forecasts, Methodology, and Metadata sheets
"""

import io
from datetime import date

from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse

from api.data_loader import get_forecasts, get_companies, get_scenario_forecasts

router = APIRouter(prefix="/api/v1", tags=["export"])


def _filter_forecasts(
    segment: str | None,
    scenario: str,
    valuation: str,
) -> "pd.DataFrame":
    """Return filtered forecast DataFrame with standardised column names.

    When *scenario* is ``"base"`` (default) the main ensemble forecasts are
    used.  For ``"conservative"`` or ``"aggressive"`` the scenario-specific
    forecasts are loaded instead.
    """
    import pandas as pd

    if scenario in ("conservative", "aggressive"):
        df = get_scenario_forecasts()
        if not df.empty:
            df = df[df["scenario"] == scenario]
    else:
        df = get_forecasts()

    if df.empty:
        return df

    if segment:
        df = df[df["segment"] == segment]

    # Select value columns based on valuation
    if valuation == "real_2020":
        rename_map = {
            "point_estimate_real_2020": "point_estimate",
            "ci80_lower": "ci80_lower",
            "ci80_upper": "ci80_upper",
            "ci95_lower": "ci95_lower",
            "ci95_upper": "ci95_upper",
        }
    else:
        rename_map = {
            "point_estimate_nominal": "point_estimate",
            "ci80_lower_nominal": "ci80_lower",
            "ci80_upper_nominal": "ci80_upper",
            "ci95_lower_nominal": "ci95_lower",
            "ci95_upper_nominal": "ci95_upper",
        }

    # For scenario data the column names may already be plain
    plain_fallback = {
        "point_estimate": "point_estimate",
        "ci80_lower": "ci80_lower",
        "ci80_upper": "ci80_upper",
        "ci95_lower": "ci95_lower",
        "ci95_upper": "ci95_upper",
    }

    for src, dst in rename_map.items():
        if src in df.columns:
            df[dst] = df[src]
        elif dst not in df.columns and dst in plain_fallback and plain_fallback[dst] in df.columns:
            pass  # already present

    base_cols = ["year", "quarter", "segment", "is_forecast"]
    value_cols = ["point_estimate", "ci80_lower", "ci80_upper", "ci95_lower", "ci95_upper"]
    cols = base_cols + value_cols

    present = [c for c in cols if c in df.columns]
    result = df[present].copy()

    # Round numeric values to 1 decimal place
    for col in value_cols:
        if col in result.columns:
            result[col] = result[col].round(1)

    # Rename columns for presentation
    result = result.rename(columns={
        "year": "Year",
        "quarter": "Quarter",
        "segment": "Segment",
        "is_forecast": "Type",
        "point_estimate": "Point Estimate ($B)",
        "ci80_lower": "CI80 Low ($B)",
        "ci80_upper": "CI80 High ($B)",
        "ci95_lower": "CI95 Low ($B)",
        "ci95_upper": "CI95 High ($B)",
    })

    # Convert boolean flag to readable label
    if "Type" in result.columns:
        result["Type"] = result["Type"].map({True: "Forecast", False: "Historical"})

    return result


def _build_filename(segment: str | None, scenario: str, ext: str) -> str:
    seg_label = segment or "all"
    return f"ai_industry_{seg_label}_{scenario}_{date.today().isoformat()}.{ext}"


# ── CSV endpoint ────────────────────────────────────────────────────────────

@router.get("/export/csv")
def export_csv(
    segment: str | None = Query(None),
    scenario: str = Query("base"),
    valuation: str = Query("nominal"),
):
    df = _filter_forecasts(segment, scenario, valuation)
    buf = io.StringIO()
    df.to_csv(buf, index=False)
    buf.seek(0)

    filename = _build_filename(segment, scenario, "csv")

    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Excel endpoint ──────────────────────────────────────────────────────────

@router.get("/export/excel")
def export_excel(
    segment: str | None = Query(None),
    scenario: str = Query("base"),
    valuation: str = Query("nominal"),
):
    import pandas as pd
    from openpyxl.styles import Alignment, Font, numbers

    fc_df = _filter_forecasts(segment, scenario, valuation)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # ── Sheet 1: Forecasts ──────────────────────────────────────────
        fc_df.to_excel(writer, sheet_name="Forecasts", index=False)

        # ── Sheet 2: Methodology ────────────────────────────────────────
        meth_df = pd.DataFrame({
            "Section": [
                "Data Sources",
                "Model",
                "Confidence Intervals",
                "Scope Normalization",
                "References",
            ],
            "Description": [
                "12 analyst firms (IDC, Gartner, McKinsey, BCG, Precedence Research, "
                "Grand View Research, MarketsandMarkets, Fortune Business Insights, "
                "Statista, CB Insights, PwC, Mordor Intelligence). "
                "EDGAR 10-K/10-Q filings for company-level attribution.",
                "ARIMA + Prophet + LightGBM inverse-RMSE-weighted ensemble. "
                "Quarterly granularity with expanding-window cross-validation.",
                "Bootstrap (1000 resamples), empirical percentiles. "
                "CI80 floor 25%, CI95 floor 40% of point estimate.",
                "Per-entry segment_scope_coefficient applied to broad-scope "
                "analyst figures to normalize to consistent segment definitions.",
                "World Bank API (GDP deflators), OECD SDMX (macro indicators), "
                "LSEG Workspace (financial data).",
            ],
        })
        meth_df.to_excel(writer, sheet_name="Methodology", index=False)

        # ── Sheet 3: Metadata ───────────────────────────────────────────
        meta_df = pd.DataFrame({
            "Field": [
                "Report Generated",
                "Segment",
                "Scenario",
                "Valuation Basis",
                "Data Vintage",
                "Tool",
            ],
            "Value": [
                date.today().isoformat(),
                segment or "All Segments",
                scenario.title(),
                "Nominal USD" if valuation == "nominal" else "Real 2020 USD",
                f"Q1 {date.today().year}",
                "AI Industry Value Estimator",
            ],
        })
        meta_df.to_excel(writer, sheet_name="Metadata", index=False)

        # ── Formatting ──────────────────────────────────────────────────
        bold_font = Font(bold=True)
        header_font = Font(bold=True, size=11)
        right_align = Alignment(horizontal="right")
        one_decimal = "0.0"

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]

            # Bold + freeze header row
            for cell in ws[1]:
                cell.font = header_font
            ws.freeze_panes = "A2"

            # Auto-width columns
            for col in ws.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value is not None:
                            max_len = max(max_len, len(str(cell.value)))
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

        # Number formatting on Forecasts sheet
        ws_fc = writer.sheets["Forecasts"]
        for row in ws_fc.iter_rows(min_row=2, max_row=ws_fc.max_row):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = one_decimal

    buf.seek(0)
    filename = _build_filename(segment, scenario, "xlsx")

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
